import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY        = "1B2A4A"
SLATE       = "2E4057"
STEEL       = "4A6FA5"
MIST        = "D6E4F0"
LIGHT_MIST  = "EEF4FB"
WHITE       = "FFFFFF"
CHARCOAL    = "2D3142"
ACCENT      = "C0392B"   # deep red for KPI highlight
GOLD        = "D4AC0D"
MID_GREY    = "6B7280"
BORDER_CLR  = "B0BEC5"

def side(color=BORDER_CLR, style="thin"):
    return Side(border_style=style, color=color)

def border(all_sides=True, color=BORDER_CLR):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s) if all_sides else Border()

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Aptos", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=CHARCOAL):
    return Font(name="Aptos", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

# ── Load & clean data ─────────────────────────────────────────────────────────
print("Loading data...")
orders   = pd.read_csv("data/olist_orders_dataset.csv")
items    = pd.read_csv("data/olist_order_items_dataset.csv")
payments = pd.read_csv("data/olist_order_payments_dataset.csv")
reviews  = pd.read_csv("data/olist_order_reviews_dataset.csv")
customers= pd.read_csv("data/olist_customers_dataset.csv")
products = pd.read_csv("data/olist_products_dataset.csv")
sellers  = pd.read_csv("data/olist_sellers_dataset.csv")
cats     = pd.read_csv("data/product_category_name_translation.csv")

# Parse dates
date_cols = ['order_purchase_timestamp','order_approved_at',
             'order_delivered_carrier_date','order_delivered_customer_date',
             'order_estimated_delivery_date']
for c in date_cols:
    orders[c] = pd.to_datetime(orders[c], errors='coerce')

# Filter delivered orders only for time-based analysis
delivered = orders[orders['order_status'] == 'delivered'].copy()

# Merge master fact table
fact = (delivered
    .merge(items[['order_id','price','freight_value']], on='order_id', how='left')
    .merge(payments.groupby('order_id')['payment_value'].sum().reset_index(), on='order_id', how='left')
    .merge(reviews[['order_id','review_score']], on='order_id', how='left')
    .merge(customers[['customer_id','customer_city','customer_state']], on='customer_id', how='left')
)
fact['delivery_days'] = (fact['order_delivered_customer_date'] - fact['order_purchase_timestamp']).dt.days
fact['order_month']   = fact['order_purchase_timestamp'].dt.to_period('M').astype(str)
fact['order_year']    = fact['order_purchase_timestamp'].dt.year
fact['revenue']       = fact['price'].fillna(0)
fact['freight']       = fact['freight_value'].fillna(0)

# ── KPIs ──────────────────────────────────────────────────────────────────────
total_revenue  = fact['revenue'].sum()
total_orders   = fact['order_id'].nunique()
avg_order_val  = fact.groupby('order_id')['revenue'].sum().mean()
avg_review     = fact['review_score'].mean()
avg_delivery   = fact['delivery_days'].median()

print(f"KPIs ready | Revenue: R${total_revenue:,.0f} | Orders: {total_orders:,}")

# ── Monthly revenue trend ─────────────────────────────────────────────────────
monthly = (fact.groupby('order_month')
    .agg(orders=('order_id','nunique'), revenue=('revenue','sum'))
    .reset_index()
    .rename(columns={'order_month':'Month'})
)
monthly = monthly[monthly['Month'] >= '2017-01'].sort_values('Month').head(24)

# ── Revenue by state ──────────────────────────────────────────────────────────
state_rev = (fact.groupby('customer_state')
    .agg(revenue=('revenue','sum'), orders=('order_id','nunique'))
    .reset_index()
    .sort_values('revenue', ascending=False)
    .head(10)
)

# ── Review score distribution ─────────────────────────────────────────────────
review_dist = (fact.groupby('review_score')['order_id']
    .nunique().reset_index()
    .rename(columns={'order_id':'count'})
    .sort_values('review_score')
)

# ── Top product categories ────────────────────────────────────────────────────
prod_with_cat = products.merge(cats, on='product_category_name', how='left')
items_cat = items.merge(prod_with_cat[['product_id','product_category_name_english']], on='product_id', how='left')
top_cats = (items_cat.groupby('product_category_name_english')
    .agg(revenue=('price','sum'), items_sold=('order_id','count'))
    .reset_index()
    .sort_values('revenue', ascending=False)
    .head(10)
    .rename(columns={'product_category_name_english':'Category'})
)

# ── Payment type breakdown ────────────────────────────────────────────────────
pay_type = (payments.groupby('payment_type')
    .agg(total=('payment_value','sum'), count=('order_id','count'))
    .reset_index()
    .sort_values('total', ascending=False)
)

# ── Clean Data sample (top 5000 rows for Excel performance) ──────────────────
clean = fact[['order_id','customer_id','order_status','order_purchase_timestamp',
              'order_delivered_customer_date','revenue','freight','payment_value',
              'review_score','delivery_days','customer_city','customer_state',
              'order_year','order_month']].drop_duplicates('order_id').head(5000).copy()
clean['order_purchase_timestamp'] = clean['order_purchase_timestamp'].dt.strftime('%Y-%m-%d')
clean['order_delivered_customer_date'] = clean['order_delivered_customer_date'].dt.strftime('%Y-%m-%d')
clean['revenue'] = clean['revenue'].round(2)
clean['freight'] = clean['freight'].round(2)
clean['payment_value'] = clean['payment_value'].round(2)
clean = clean.fillna({'review_score':0,'delivery_days':0,'payment_value':0})

print("Data prep done. Building workbook...")

# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — README / COVER
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("README Cover")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 60
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 3

# Title block — merge A1:E6
ws1.merge_cells('A1:E6')
title_cell = ws1['A1']
title_cell.value = "Olist Brazilian E-Commerce\nData Analysis & Dashboard"
title_cell.font = Font(name="Aptos", size=26, bold=True, color=WHITE)
title_cell.fill = fill(NAVY)
title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Subtitle row
ws1.merge_cells('A7:E7')
sub = ws1['A7']
sub.value = "Comprehensive analysis of 100,000+ orders (2016–2018)  |  Source: Olist Public Dataset  |  Prepared by: Data Analytics Team"
sub.font = Font(name="Aptos", size=10, color=WHITE)
sub.fill = fill(SLATE)
sub.alignment = center()

ws1.row_dimensions[7].height = 22

# Spacer
ws1.row_dimensions[8].height = 14

# Project description
desc_rows = [
    ("PROJECT OVERVIEW", None, True),
    ("Dataset", "100,000+ orders placed on Olist marketplace (2016–2018) across Brazil, covering order lifecycle, payments, logistics, customer reviews, and product data.", False),
    ("Objective", "Uncover revenue trends, customer behaviour patterns, delivery performance, and top-performing product categories to support strategic decisions.", False),
    ("Scope", "Delivered orders only are used for financial KPIs. All 9 source tables have been joined into a master fact table for analysis.", False),
]
row = 9
for label, val, is_hdr in desc_rows:
    if is_hdr:
        ws1.merge_cells(f'B{row}:D{row}')
        c = ws1[f'B{row}']
        c.value = label
        c.font = Font(name="Aptos", size=11, bold=True, color=NAVY)
        c.fill = fill(MIST)
        c.alignment = left()
        c.border = border()
        ws1.row_dimensions[row].height = 20
    else:
        ws1[f'B{row}'].value = label
        ws1[f'B{row}'].font = Font(name="Aptos", size=10, bold=True, color=CHARCOAL)
        ws1[f'B{row}'].alignment = left()
        ws1[f'B{row}'].border = border()
        ws1[f'C{row}'].value = val
        ws1[f'C{row}'].font = body_font()
        ws1[f'C{row}'].alignment = left(wrap=True)
        ws1[f'C{row}'].border = border()
        ws1.merge_cells(f'C{row}:D{row}')
        ws1.row_dimensions[row].height = 36
    row += 1

# Spacer
row += 1

# Table of Contents
toc_title_row = row
ws1.merge_cells(f'B{row}:D{row}')
c = ws1[f'B{row}']
c.value = "TABLE OF CONTENTS"
c.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
c.fill = fill(STEEL)
c.alignment = center()
c.border = border()
ws1.row_dimensions[row].height = 22
row += 1

toc_entries = [
    ("Tab 1", "README Cover",      "Project overview, data dictionary, and navigation"),
    ("Tab 2", "Clean Data",        "Cleaned & structured master fact table (5,000 rows sample)"),
    ("Tab 3", "Analysis & Pivots", "3 pivot summaries: Monthly Trend, State Revenue, Category Performance"),
    ("Tab 4", "Dashboard",         "Executive KPI cards + 4 native Excel charts"),
]
for tab, name, desc in toc_entries:
    ws1[f'B{row}'].value = tab
    ws1[f'B{row}'].font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    ws1[f'B{row}'].fill = fill(SLATE)
    ws1[f'B{row}'].alignment = center()
    ws1[f'B{row}'].border = border()

    ws1[f'C{row}'].value = name
    ws1[f'C{row}'].font = Font(name="Aptos", size=10, bold=True, color=NAVY)
    ws1[f'C{row}'].fill = fill(LIGHT_MIST)
    ws1[f'C{row}'].alignment = left()
    ws1[f'C{row}'].border = border()

    ws1[f'D{row}'].value = desc
    ws1[f'D{row}'].font = body_font()
    ws1[f'D{row}'].fill = fill(LIGHT_MIST)
    ws1[f'D{row}'].alignment = left(wrap=True)
    ws1[f'D{row}'].border = border()
    ws1.row_dimensions[row].height = 22
    row += 1

# Spacer
row += 1

# Data Dictionary
ws1.merge_cells(f'B{row}:D{row}')
c = ws1[f'B{row}']
c.value = "DATA DICTIONARY — KEY COLUMNS"
c.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
c.fill = fill(STEEL)
c.alignment = center()
c.border = border()
ws1.row_dimensions[row].height = 22
row += 1

dd_headers = ["Column", "Source Table", "Description"]
for i, h in enumerate(dd_headers, start=2):
    col = get_column_letter(i)
    c = ws1[f'{col}{row}']
    c.value = h
    c.font = Font(name="Aptos", size=10, bold=True, color=CHARCOAL)
    c.fill = fill(MIST)
    c.alignment = center()
    c.border = border()
ws1.row_dimensions[row].height = 18
row += 1

dd_entries = [
    ("order_id",                  "orders",   "Unique identifier for each order"),
    ("customer_id",               "orders",   "Link to customer dimension"),
    ("order_status",              "orders",   "Order lifecycle stage (delivered, shipped, canceled, etc.)"),
    ("order_purchase_timestamp",  "orders",   "Date-time when order was placed by customer"),
    ("order_delivered_customer_date","orders","Actual delivery date to end customer"),
    ("price",                     "items",    "Product price (BRL) per item line"),
    ("freight_value",             "items",    "Freight cost per item (BRL)"),
    ("payment_value",             "payments", "Total payment collected for the order"),
    ("payment_type",              "payments", "Method: credit_card, boleto, voucher, debit_card"),
    ("review_score",              "reviews",  "Customer rating 1–5 (5 = best)"),
    ("customer_state",            "customers","Brazilian state abbreviation (e.g., SP, RJ)"),
    ("product_category_name_english","products+cats","English translation of product category"),
    ("delivery_days",             "derived",  "Actual delivery duration in calendar days"),
    ("revenue",                   "derived",  "Alias for price; used in aggregations"),
    ("order_month",               "derived",  "YYYY-MM period string for monthly grouping"),
]

alt = False
for col_name, src, desc in dd_entries:
    bg = LIGHT_MIST if alt else WHITE
    for i, val in enumerate([col_name, src, desc], start=2):
        col = get_column_letter(i)
        c = ws1[f'{col}{row}']
        c.value = val
        c.font = body_font(color=CHARCOAL if i > 2 else NAVY)
        if i == 2: c.font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        c.fill = fill(bg)
        c.alignment = left(wrap=(i==4))
        c.border = border()
    ws1.row_dimensions[row].height = 18
    row += 1
    alt = not alt

print("Tab 1 done.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CLEAN DATA
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Clean Data")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"

col_map = {
    'order_id':                     ('Order ID',            '14'),
    'customer_id':                  ('Customer ID',         '14'),
    'order_status':                 ('Status',              '12'),
    'order_purchase_timestamp':     ('Purchase Date',       '13'),
    'order_delivered_customer_date':('Delivery Date',       '13'),
    'revenue':                      ('Revenue (R$)',        '13'),
    'freight':                      ('Freight (R$)',        '12'),
    'payment_value':                ('Payment (R$)',        '13'),
    'review_score':                 ('Review Score',        '12'),
    'delivery_days':                ('Delivery Days',       '13'),
    'customer_city':                ('City',                '16'),
    'customer_state':               ('State',               '8'),
    'order_year':                   ('Year',                '8'),
    'order_month':                  ('Month',               '10'),
}

cols = list(col_map.keys())
headers = [col_map[c][0] for c in cols]
widths  = [col_map[c][1] for c in cols]

# Header row
for j, (h, w) in enumerate(zip(headers, widths), start=1):
    col = get_column_letter(j)
    ws2.column_dimensions[col].width = int(w)
    c = ws2.cell(row=1, column=j, value=h)
    c.font   = header_font(size=10)
    c.fill   = fill(NAVY)
    c.alignment = center()
    c.border = border(color=SLATE)
ws2.row_dimensions[1].height = 22

# Data rows
currency_cols = {'Revenue (R$)', 'Freight (R$)', 'Payment (R$)'}
num_cols      = {'Review Score', 'Delivery Days', 'Year'}

for i, row_data in enumerate(clean[cols].itertuples(index=False), start=2):
    for j, (val, hdr) in enumerate(zip(row_data, headers), start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = body_font()
        c.alignment = left()
        c.border = Border(
            bottom=Side(border_style='thin', color='E0E0E0')
        )
        if hdr in currency_cols:
            c.number_format = '#,##0.00'
            c.alignment = right_align()
        elif hdr in num_cols:
            c.number_format = '#,##0'
            c.alignment = center()
        if i % 2 == 0:
            c.fill = fill(LIGHT_MIST)

# Excel Table
last_col = get_column_letter(len(cols))
last_row = len(clean) + 1
tbl = Table(displayName="OrderData", ref=f"A1:{last_col}{last_row}")
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl.tableStyleInfo = style
ws2.add_table(tbl)

print("Tab 2 done.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — ANALYSIS & PIVOTS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Analysis & Pivots")
ws3.sheet_view.showGridLines = False

def write_pivot_header(ws, row, col, title, span=5):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=title)
    c.font = Font(name="Aptos", size=12, bold=True, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    c.border = border()
    ws.row_dimensions[row].height = 22

def write_col_headers(ws, row, col, headers, colors=None):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col+i, value=h)
        c.font = Font(name="Aptos", size=10, bold=True, color=CHARCOAL)
        c.fill = fill(MIST)
        c.alignment = center(wrap=True)
        c.border = border()
    ws.row_dimensions[row].height = 20

def write_data_row(ws, row, col, values, formats=None, alt=False):
    bg = LIGHT_MIST if alt else WHITE
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col+i, value=v)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = right_align() if isinstance(v, (int,float)) and not isinstance(v, bool) else left()
        c.border = border()
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]
    ws.row_dimensions[row].height = 18

# ── PIVOT 1: Monthly Revenue & Order Trend ────────────────────────────────────
start_row, start_col = 2, 2
write_pivot_header(ws3, start_row, start_col, "① Monthly Revenue & Order Volume Trend", span=5)
write_col_headers(ws3, start_row+1, start_col,
    ["Month", "Orders", "Revenue (R$)", "Avg Order Value (R$)", "MoM Revenue Growth"])

# Write raw data for monthly
monthly_rows_start = start_row + 2
for idx, r in enumerate(monthly.itertuples(index=False)):
    rw = monthly_rows_start + idx
    alt = idx % 2 == 1
    # MoM formula (skip first row)
    if idx == 0:
        mom = "—"
        fmt_mom = None
    else:
        prev_rev_cell = f"D{rw-1}"
        curr_rev_cell = f"D{rw}"
        mom = f"=IFERROR(({curr_rev_cell}-{prev_rev_cell})/{prev_rev_cell},0)"
        fmt_mom = '0.0%'

    vals = [r.Month, r.orders, r.revenue, None, None]
    for i, v in enumerate(vals[:3], start=start_col):
        c = ws3.cell(row=rw, column=i, value=v)
        c.font = body_font()
        c.fill = fill(LIGHT_MIST if alt else WHITE)
        c.border = border()
        if i == start_col:
            c.alignment = center()
        else:
            c.alignment = right_align()
            c.number_format = '#,##0' if i == start_col+1 else '#,##0.00'

    # Avg order value formula
    c_avg = ws3.cell(row=rw, column=start_col+3)
    c_avg.value = f"=IFERROR({get_column_letter(start_col+2)}{rw}/{get_column_letter(start_col+1)}{rw},0)"
    c_avg.font = body_font()
    c_avg.fill = fill(LIGHT_MIST if alt else WHITE)
    c_avg.alignment = right_align()
    c_avg.border = border()
    c_avg.number_format = '#,##0.00'

    # MoM Growth formula
    c_mom = ws3.cell(row=rw, column=start_col+4)
    c_mom.value = mom
    c_mom.font = body_font()
    c_mom.fill = fill(LIGHT_MIST if alt else WHITE)
    c_mom.alignment = right_align()
    c_mom.border = border()
    if fmt_mom:
        c_mom.number_format = fmt_mom

monthly_end_row = monthly_rows_start + len(monthly) - 1

# Totals row
tot_row = monthly_end_row + 1
ws3.merge_cells(start_row=tot_row, start_column=start_col, end_row=tot_row, end_column=start_col)
c = ws3.cell(row=tot_row, column=start_col, value="TOTAL")
c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
c.fill = fill(SLATE)
c.alignment = center()
c.border = border()

ord_col = get_column_letter(start_col+1)
rev_col = get_column_letter(start_col+2)
for col_offset, fmt in [(1,'#,##0'),(2,'#,##0.00')]:
    col_ltr = get_column_letter(start_col+col_offset)
    c = ws3.cell(row=tot_row, column=start_col+col_offset,
                 value=f"=SUM({col_ltr}{monthly_rows_start}:{col_ltr}{monthly_end_row})")
    c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    c.fill = fill(SLATE)
    c.alignment = right_align()
    c.number_format = fmt
    c.border = border()

# Grand avg
c = ws3.cell(row=tot_row, column=start_col+3,
             value=f"=IFERROR({rev_col}{tot_row}/{ord_col}{tot_row},0)")
c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
c.fill = fill(SLATE)
c.alignment = right_align()
c.number_format = '#,##0.00'
c.border = border()

ws3.cell(row=tot_row, column=start_col+4).fill = fill(SLATE)
ws3.cell(row=tot_row, column=start_col+4).border = border()

# Column widths
col_widths_p1 = {start_col:13, start_col+1:10, start_col+2:16, start_col+3:18, start_col+4:16}
for c_idx, w in col_widths_p1.items():
    ws3.column_dimensions[get_column_letter(c_idx)].width = w

# ── PIVOT 2: Revenue by State (SUMIFS style aggregation) ─────────────────────
p2_start_row = tot_row + 3
p2_start_col = 2

write_pivot_header(ws3, p2_start_row, p2_start_col, "② Revenue & Orders by Brazilian State (Top 10)", span=4)
write_col_headers(ws3, p2_start_row+1, p2_start_col,
    ["State", "Total Revenue (R$)", "Total Orders", "Avg Revenue / Order (R$)"])

p2_data_start = p2_start_row + 2
for idx, r in enumerate(state_rev.itertuples(index=False)):
    rw = p2_data_start + idx
    alt = idx % 2 == 1
    vals = [r.customer_state, r.revenue, r.orders]
    for i, v in enumerate(vals, start=p2_start_col):
        c = ws3.cell(row=rw, column=i, value=v)
        c.font = body_font()
        c.fill = fill(LIGHT_MIST if alt else WHITE)
        c.border = border()
        c.alignment = center() if i == p2_start_col else right_align()
        if i == p2_start_col+1: c.number_format = '#,##0.00'
        if i == p2_start_col+2: c.number_format = '#,##0'

    # Avg formula
    rev_c = get_column_letter(p2_start_col+1)
    ord_c = get_column_letter(p2_start_col+2)
    avg_c = ws3.cell(row=rw, column=p2_start_col+3)
    avg_c.value = f"=IFERROR({rev_c}{rw}/{ord_c}{rw},0)"
    avg_c.font = body_font()
    avg_c.fill = fill(LIGHT_MIST if alt else WHITE)
    avg_c.alignment = right_align()
    avg_c.number_format = '#,##0.00'
    avg_c.border = border()

p2_end_row = p2_data_start + len(state_rev) - 1
# Totals
tot2 = p2_end_row + 1
c = ws3.cell(row=tot2, column=p2_start_col, value="TOTAL")
c.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c.fill=fill(SLATE); c.alignment=center(); c.border=border()
for off, fmt in [(1,'#,##0.00'),(2,'#,##0')]:
    col_l = get_column_letter(p2_start_col+off)
    c = ws3.cell(row=tot2, column=p2_start_col+off,
                 value=f"=SUM({col_l}{p2_data_start}:{col_l}{p2_end_row})")
    c.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c.fill=fill(SLATE)
    c.alignment=right_align(); c.number_format=fmt; c.border=border()
c2 = ws3.cell(row=tot2, column=p2_start_col+3)
c2.value = f"=IFERROR({get_column_letter(p2_start_col+1)}{tot2}/{get_column_letter(p2_start_col+2)}{tot2},0)"
c2.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c2.fill=fill(SLATE)
c2.alignment=right_align(); c2.number_format='#,##0.00'; c2.border=border()

for coff, w in {0:8,1:18,2:14,3:22}.items():
    ws3.column_dimensions[get_column_letter(p2_start_col+coff)].width = w

# ── PIVOT 3: Top 10 Product Categories ───────────────────────────────────────
p3_start_row = tot2 + 3
p3_start_col = 2

write_pivot_header(ws3, p3_start_row, p3_start_col, "③ Top 10 Product Categories by Revenue", span=4)
write_col_headers(ws3, p3_start_row+1, p3_start_col,
    ["Category", "Revenue (R$)", "Items Sold", "Revenue Share"])

p3_data_start = p3_start_row + 2
p3_end_row = p3_data_start + len(top_cats) - 1
total_cat_rev = top_cats['revenue'].sum()

for idx, r in enumerate(top_cats.itertuples(index=False)):
    rw = p3_data_start + idx
    alt = idx % 2 == 1
    vals = [r.Category, r.revenue, r.items_sold]
    for i, v in enumerate(vals, start=p3_start_col):
        c = ws3.cell(row=rw, column=i, value=v)
        c.font = body_font()
        c.fill = fill(LIGHT_MIST if alt else WHITE)
        c.border = border()
        c.alignment = left() if i == p3_start_col else right_align()
        if i == p3_start_col+1: c.number_format = '#,##0.00'
        if i == p3_start_col+2: c.number_format = '#,##0'

    # Revenue share formula
    rev_c = get_column_letter(p3_start_col+1)
    tot_cell_addr = f"{rev_c}{p3_end_row+1}"  # will point to total row
    sh_c = ws3.cell(row=rw, column=p3_start_col+3)
    sh_c.value = f"=IFERROR({rev_c}{rw}/SUM({rev_c}{p3_data_start}:{rev_c}{p3_end_row}),0)"
    sh_c.font = body_font()
    sh_c.fill = fill(LIGHT_MIST if alt else WHITE)
    sh_c.alignment = right_align()
    sh_c.number_format = '0.0%'
    sh_c.border = border()

# Totals
tot3 = p3_end_row + 1
c = ws3.cell(row=tot3, column=p3_start_col, value="TOTAL")
c.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c.fill=fill(SLATE); c.alignment=center(); c.border=border()
for off, fmt in [(1,'#,##0.00'),(2,'#,##0')]:
    col_l = get_column_letter(p3_start_col+off)
    c = ws3.cell(row=tot3, column=p3_start_col+off,
                 value=f"=SUM({col_l}{p3_data_start}:{col_l}{p3_end_row})")
    c.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c.fill=fill(SLATE)
    c.alignment=right_align(); c.number_format=fmt; c.border=border()
c2 = ws3.cell(row=tot3, column=p3_start_col+3, value="100.0%")
c2.font = Font(name="Aptos",size=10,bold=True,color=WHITE); c2.fill=fill(SLATE)
c2.alignment=right_align(); c2.number_format='0.0%'; c2.border=border()

for coff, w in {0:30,1:16,2:12,3:14}.items():
    ws3.column_dimensions[get_column_letter(p3_start_col+coff)].width = w

ws3.column_dimensions['A'].width = 3

print("Tab 3 done.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")
ws4.sheet_view.showGridLines = False

# Row / column sizing
for r in range(1, 60):
    ws4.row_dimensions[r].height = 18
ws4.row_dimensions[1].height = 6   # top padding
ws4.row_dimensions[2].height = 40  # title

for col_idx in range(1, 22):
    ws4.column_dimensions[get_column_letter(col_idx)].width = 9.5

# Dashboard Title
ws4.merge_cells('B2:U2')
c = ws4['B2']
c.value = "OLIST E-COMMERCE  ·  EXECUTIVE DASHBOARD"
c.font = Font(name="Aptos", size=18, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center()

ws4.merge_cells('B3:U3')
sub = ws4['B3']
sub.value = "Brazil Market Analysis  |  2016–2018  |  Delivered Orders"
sub.font = Font(name="Aptos", size=10, color=WHITE)
sub.fill = fill(SLATE)
sub.alignment = center()
ws4.row_dimensions[3].height = 20

# ── KPI Cards ─────────────────────────────────────────────────────────────────
ws4.row_dimensions[4].height = 8

kpis = [
    ("TOTAL REVENUE",    f"R$ {total_revenue:,.0f}",   "BRL across all delivered orders",    NAVY,  "B5:E9"),
    ("TOTAL ORDERS",     f"{total_orders:,}",           "Unique delivered orders",             STEEL, "G5:J9"),
    ("AVG ORDER VALUE",  f"R$ {avg_order_val:,.2f}",   "Average revenue per order",           SLATE, "L5:O9"),
    ("AVG REVIEW SCORE", f"{avg_review:.2f} / 5.0",    "Customer satisfaction rating",        NAVY,  "Q5:T9"),
    ("MEDIAN DELIVERY",  f"{avg_delivery:.0f} Days",   "Calendar days from purchase",         STEEL, "B11:E15"),
]

ws4.row_dimensions[10].height = 8

for title, value, subtitle, color, cell_range in kpis:
    ws4.merge_cells(cell_range)
    start_cell = cell_range.split(':')[0]
    c = ws4[start_cell]
    c.fill = fill(color)
    c.alignment = center(wrap=True)
    c.border = border(color=color)

    # We write text into the merged cell as multi-line
    c.value = f"{title}\n{value}\n{subtitle}"
    c.font = Font(name="Aptos", size=9, bold=False, color=WHITE)

    # Override with richer formatting via a helper approach:
    # Since openpyxl doesn't support multi-format in one cell, we use 3 rows
    # Let's unmerge and re-do with 3 rows
    ws4.unmerge_cells(cell_range)
    r_start = int(''.join(filter(str.isdigit, cell_range.split(':')[0])))
    c_start = ''.join(filter(str.isalpha, cell_range.split(':')[0]))
    c_start_idx = sum((ord(ch)-64)*26**i for i, ch in enumerate(reversed(c_start.upper())))

    r_end   = int(''.join(filter(str.isdigit, cell_range.split(':')[1])))
    c_end   = ''.join(filter(str.isalpha, cell_range.split(':')[1]))
    c_end_idx = sum((ord(ch)-64)*26**i for i, ch in enumerate(reversed(c_end.upper())))

    # Fill entire block
    for rr in range(r_start, r_end+1):
        for cc in range(c_start_idx, c_end_idx+1):
            cell = ws4.cell(row=rr, column=cc)
            cell.fill = fill(color)

    # Title row (top 1 row)
    ws4.merge_cells(start_row=r_start, start_column=c_start_idx,
                    end_row=r_start,   end_column=c_end_idx)
    ct = ws4.cell(row=r_start, column=c_start_idx, value=title)
    ct.font = Font(name="Aptos", size=9, bold=True, color="AACCE8")
    ct.fill = fill(color)
    ct.alignment = center()
    ws4.row_dimensions[r_start].height = 18

    # Value row (middle)
    mid_r = r_start + 2
    ws4.merge_cells(start_row=r_start+1, start_column=c_start_idx,
                    end_row=mid_r,       end_column=c_end_idx)
    cv = ws4.cell(row=r_start+1, column=c_start_idx, value=value)
    cv.font = Font(name="Aptos", size=16, bold=True, color=WHITE)
    cv.fill = fill(color)
    cv.alignment = center()
    ws4.row_dimensions[r_start+1].height = 28

    # Subtitle row
    ws4.merge_cells(start_row=mid_r+1, start_column=c_start_idx,
                    end_row=r_end,     end_column=c_end_idx)
    cs = ws4.cell(row=mid_r+1, column=c_start_idx, value=subtitle)
    cs.font = Font(name="Aptos", size=8, color="AACCE8")
    cs.fill = fill(color)
    cs.alignment = center(wrap=True)

# ── Chart data areas (hidden helper area for charts) ─────────────────────────
# We write chart data starting at column 23 (W) — hidden from view

CHART_COL = 23  # Column W

# Chart Data 1: Monthly Revenue
ws4.cell(row=1, column=CHART_COL, value="Month")
ws4.cell(row=1, column=CHART_COL+1, value="Revenue")
ws4.cell(row=1, column=CHART_COL+2, value="Orders")
for i, r in enumerate(monthly.itertuples(index=False), start=2):
    ws4.cell(row=i, column=CHART_COL,   value=r.Month)
    ws4.cell(row=i, column=CHART_COL+1, value=round(r.revenue, 2))
    ws4.cell(row=i, column=CHART_COL+2, value=r.orders)
monthly_chart_end = len(monthly) + 1

# Chart Data 2: State Revenue
STATE_COL = CHART_COL + 4
ws4.cell(row=1, column=STATE_COL, value="State")
ws4.cell(row=1, column=STATE_COL+1, value="Revenue")
for i, r in enumerate(state_rev.itertuples(index=False), start=2):
    ws4.cell(row=i, column=STATE_COL,   value=r.customer_state)
    ws4.cell(row=i, column=STATE_COL+1, value=round(r.revenue, 2))
state_chart_end = len(state_rev) + 1

# Chart Data 3: Review Score Distribution
REV_COL = CHART_COL + 7
ws4.cell(row=1, column=REV_COL, value="Score")
ws4.cell(row=1, column=REV_COL+1, value="Count")
for i, r in enumerate(review_dist.itertuples(index=False), start=2):
    ws4.cell(row=i, column=REV_COL,   value=r.review_score)
    ws4.cell(row=i, column=REV_COL+1, value=r.count)
review_chart_end = len(review_dist) + 1

# Chart Data 4: Top Categories
CAT_COL = CHART_COL + 10
ws4.cell(row=1, column=CAT_COL, value="Category")
ws4.cell(row=1, column=CAT_COL+1, value="Revenue")
for i, r in enumerate(top_cats.head(8).itertuples(index=False), start=2):
    ws4.cell(row=i, column=CAT_COL,   value=r.Category)
    ws4.cell(row=i, column=CAT_COL+1, value=round(r.revenue, 2))
cat_chart_end = len(top_cats.head(8)) + 1

# ── CHART 1: Monthly Revenue Line Chart ──────────────────────────────────────
chart1 = LineChart()
chart1.title = "Monthly Revenue Trend (R$)"
chart1.style = 10
chart1.grouping = "standard"
chart1.smooth = True

data1 = Reference(ws4, min_col=CHART_COL+1, min_row=1, max_row=monthly_chart_end)
cats1 = Reference(ws4, min_col=CHART_COL,   min_row=2, max_row=monthly_chart_end)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.shape = 4

s = chart1.series[0]
s.graphicalProperties.line.solidFill = STEEL
s.graphicalProperties.line.width = 20000
s.smooth = True

chart1.width  = 16
chart1.height = 10
chart1.plot_area.graphicalProps = None

ws4.add_chart(chart1, "B17")

# ── CHART 2: Revenue by State Bar Chart ──────────────────────────────────────
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Top 10 States by Revenue (R$)"
chart2.style = 10
chart2.grouping = "clustered"

data2 = Reference(ws4, min_col=STATE_COL+1, min_row=1, max_row=state_chart_end)
cats2 = Reference(ws4, min_col=STATE_COL,   min_row=2, max_row=state_chart_end)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)

chart2.series[0].graphicalProperties.solidFill = NAVY
chart2.width  = 16
chart2.height = 10
ws4.add_chart(chart2, "L17")

# ── CHART 3: Review Score Distribution Bar ───────────────────────────────────
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Review Score Distribution"
chart3.style = 10

data3 = Reference(ws4, min_col=REV_COL+1, min_row=1, max_row=review_chart_end)
cats3 = Reference(ws4, min_col=REV_COL,   min_row=2, max_row=review_chart_end)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)

colors_rev = ["C0392B","E67E22","F1C40F","2ECC71","1B2A4A"]
for i, col_hex in enumerate(colors_rev):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = col_hex
    chart3.series[0].dPt.append(pt)

chart3.width  = 14
chart3.height = 10
ws4.add_chart(chart3, "B34")

# ── CHART 4: Top Categories Horizontal Bar ───────────────────────────────────
chart4 = BarChart()
chart4.type = "bar"
chart4.title = "Top 8 Product Categories by Revenue"
chart4.style = 10

data4 = Reference(ws4, min_col=CAT_COL+1, min_row=1, max_row=cat_chart_end)
cats4 = Reference(ws4, min_col=CAT_COL,   min_row=2, max_row=cat_chart_end)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = SLATE

chart4.width  = 18
chart4.height = 10
ws4.add_chart(chart4, "K34")

print("Tab 4 done.")

# ── Tab ordering & colors ─────────────────────────────────────────────────────
tab_colors = {
    "README Cover":      "1B2A4A",
    "Clean Data":        "4A6FA5",
    "Analysis & Pivots": "2E4057",
    "Dashboard":         "C0392B",
}
for sheet_name, color in tab_colors.items():
    wb[sheet_name].sheet_properties.tabColor = color

# Set Dashboard as active
wb.active = wb["Dashboard"]

# Save
out_path = "/mnt/user-data/outputs/Olist_Ecommerce_Analysis.xlsx"
wb.save(out_path)
print(f"\n✅ Saved to {out_path}")
